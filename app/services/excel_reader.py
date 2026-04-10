import openpyxl
from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime
from app.config import EXCEL_FILES, EXCEL_DIR

class ExcelReader:
    """Service to read Excel files and extract data"""

    @staticmethod
    def _resolve_file_path(file_name: str, file_path: Optional[Path] = None) -> Path:
        """Resolve effective file path, allowing override during uploads."""
        if file_path is not None:
            return Path(file_path)
        return EXCEL_DIR / file_name

    @staticmethod
    def _get_sheet(workbook, config: Dict[str, Any]):
        """Get worksheet by configured name with optional aliases."""
        candidates: List[str] = [config.get("sheet_name", "")]
        candidates.extend(config.get("sheet_aliases", []))
        for name in candidates:
            if name and name in workbook.sheetnames:
                return workbook[name]
        raise ValueError(f"Sheet not found. Tried: {candidates}")

    @staticmethod
    def _unique_headers(worksheet, header_row: int, start_col: int, end_col: int) -> List[str]:
        headers: List[str] = []
        seen = set()
        for col_idx in range(start_col, end_col + 1):
            value = worksheet.cell(row=header_row, column=col_idx).value
            header = str(value).strip() if value is not None and str(value).strip() else f"Colonne_{col_idx}"
            base = header
            suffix = 2
            while header in seen:
                header = f"{base}_{suffix}"
                suffix += 1
            seen.add(header)
            headers.append(header)
        return headers

    @staticmethod
    def _row_title(worksheet, row_idx: int, start_col: int, end_col: int) -> str:
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_row <= row_idx <= merged_range.max_row:
                if merged_range.max_col < start_col or merged_range.min_col > end_col:
                    continue
                top_left = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                if top_left is not None and str(top_left).strip():
                    return str(top_left).strip()

        for col_idx in range(start_col, end_col + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is not None and str(value).strip():
                return str(value).strip()

        return "Sans section"

    @staticmethod
    def _extract_rect_table(
        worksheet,
        *,
        header_row: int,
        data_start_row: int,
        data_end_row: int,
        start_col: int,
        end_col: int,
        section_title_row: Optional[int] = None,
    ) -> Dict[str, Any]:
        headers = ExcelReader._unique_headers(worksheet, header_row, start_col, end_col)

        section_title = None
        if section_title_row is not None:
            section_title = ExcelReader._row_title(worksheet, section_title_row, start_col, end_col)

        records = []
        for row_idx in range(data_start_row, data_end_row + 1):
            row_values = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(start_col, end_col + 1)]
            if all(value is None or str(value).strip() == "" for value in row_values):
                continue

            row_record: Dict[str, Any] = {}
            if section_title is not None:
                row_record["Section"] = section_title

            for idx, header in enumerate(headers):
                value = row_values[idx]
                row_record[header] = "" if value is None else value

            records.append(row_record)

        columns = ["Section", *headers] if section_title is not None else headers
        return {"columns": columns, "data": records}
    
    @staticmethod
    def get_fiche_consommation(file_path: Optional[Path] = None) -> Dict[str, Any]:
        """Extract fiche data from A5:G36 with headers on row 5."""
        config = EXCEL_FILES["fiche_consommation"]
        source_file = ExcelReader._resolve_file_path(config["file_name"], file_path)
        if not source_file.exists():
            raise FileNotFoundError(f"File not found: {source_file}")

        workbook = openpyxl.load_workbook(source_file, data_only=True)
        worksheet = workbook[config["sheet_name"]]

        headers = []
        for col_idx in range(1, 8):  # A..G
            value = worksheet.cell(row=5, column=col_idx).value
            if value is None or str(value).strip() == "":
                headers.append(f"Colonne_{col_idx}")
            else:
                headers.append(str(value).strip())

        records = []
        for row_idx in range(6, 37):
            row_values = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 8)]
            if all(value is None for value in row_values):
                continue

            row_record = {}
            for idx, header in enumerate(headers):
                cell_value = row_values[idx]
                row_record[header] = "" if cell_value is None else cell_value
            records.append(row_record)

        workbook.close()

        return {
            "name": "Fiche Consommation Journalière",
            "data": records,
            "columns": headers,
            "row_count": len(records),
            "source_range": "A5:G36",
            "last_updated": datetime.now().isoformat()
        }
    
    @staticmethod
    def get_calcul_viande(file_path: Optional[Path] = None) -> Dict[str, Any]:
        """Extract calcul viande from A4:P22 with category title rows."""
        config = EXCEL_FILES["calcul_viande"]
        source_file = ExcelReader._resolve_file_path(config["file_name"], file_path)
        if not source_file.exists():
            raise FileNotFoundError(f"File not found: {source_file}")

        workbook = openpyxl.load_workbook(source_file, data_only=True)
        worksheet = workbook[config["sheet_name"]]

        # A4:P4 headers
        headers = []
        seen_headers = set()
        for col_idx in range(1, 17):  # A..P
            value = worksheet.cell(row=4, column=col_idx).value
            header = str(value).strip() if value is not None and str(value).strip() else f"Colonne_{col_idx}"

            # Ensure unique header names
            base_header = header
            suffix = 2
            while header in seen_headers:
                header = f"{base_header}_{suffix}"
                suffix += 1

            seen_headers.add(header)
            headers.append(header)

        category_rows = {5, 13, 15, 17, 19}

        def get_row_category_title(row_idx: int) -> str:
            # If a row is a merged title row, the title is often stored in the top-left cell
            # of the merged range. We recover that value first.
            for merged_range in worksheet.merged_cells.ranges:
                if merged_range.min_row <= row_idx <= merged_range.max_row:
                    top_left = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
                    if top_left is not None and str(top_left).strip():
                        return str(top_left).strip()

            # Fallback: first non-empty cell in A..P for this row
            for col_idx in range(1, 17):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None and str(cell_value).strip():
                    return str(cell_value).strip()

            return "Sans catégorie"

        current_category = "Sans catégorie"
        records = []

        for row_idx in range(5, 23):  # rows 5..22
            if row_idx in category_rows:
                current_category = get_row_category_title(row_idx)
                continue

            row_values = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 17)]
            if all(value is None for value in row_values):
                continue

            row_record = {"Categorie": current_category}
            for idx, header in enumerate(headers):
                value = row_values[idx]
                row_record[header] = "" if value is None else value
            records.append(row_record)

        workbook.close()

        return {
            "name": "Calculateur Ing Viande - Conso Journalière",
            "data": records,
            "columns": ["Categorie", *headers],
            "row_count": len(records),
            "source_range": "A4:P22 (row 4 headers, rows 5/13/15/17/19 as category titles)",
            "last_updated": datetime.now().isoformat()
        }
    
    @staticmethod
    def get_emballage_synthese(file_path: Optional[Path] = None) -> Dict[str, Any]:
        """Extract emballage synthèse from A5:F14 (including total row)."""
        config = EXCEL_FILES["emballage_synthese"]
        source_file = ExcelReader._resolve_file_path(config["file_name"], file_path)
        if not source_file.exists():
            raise FileNotFoundError(f"File not found: {source_file}")

        workbook = openpyxl.load_workbook(source_file, data_only=True)
        worksheet = workbook[config["sheet_name"]]

        headers = []
        seen_headers = set()
        for col_idx in range(1, 7):  # A..F
            value = worksheet.cell(row=5, column=col_idx).value
            header = str(value).strip() if value is not None and str(value).strip() else f"Colonne_{col_idx}"

            base_header = header
            suffix = 2
            while header in seen_headers:
                header = f"{base_header}_{suffix}"
                suffix += 1

            seen_headers.add(header)
            headers.append(header)

        records = []
        for row_idx in range(6, 15):  # rows 6..14 (14 is total)
            row_values = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 7)]
            if all(value is None for value in row_values):
                continue

            row_record = {}
            for idx, header in enumerate(headers):
                value = row_values[idx]
                row_record[header] = "" if value is None else value
            records.append(row_record)

        workbook.close()

        return {
            "name": "Calculateur Emballage - Synthèse Totaux",
            "data": records,
            "columns": headers,
            "row_count": len(records),
            "source_range": "A5:F14",
            "last_updated": datetime.now().isoformat()
        }

    @staticmethod
    def get_fiche_emb_ingredient(file_path: Optional[Path] = None) -> Dict[str, Any]:
        """Extract Controle Emb/Ingr from A3:H3 headers, title in row 4, data rows 5..30."""
        config = EXCEL_FILES["fiche_emb_ingredient"]
        source_file = ExcelReader._resolve_file_path(config["file_name"], file_path)
        if not source_file.exists():
            raise FileNotFoundError(f"File not found: {source_file}")

        workbook = openpyxl.load_workbook(source_file, data_only=True)
        worksheet = ExcelReader._get_sheet(workbook, config)

        extracted = ExcelReader._extract_rect_table(
            worksheet,
            header_row=3,
            data_start_row=5,
            data_end_row=30,
            start_col=1,
            end_col=8,
            section_title_row=4,
        )

        workbook.close()
        return {
            "name": "Contrôle Emballage et Ingrédients",
            "data": extracted["data"],
            "columns": extracted["columns"],
            "row_count": len(extracted["data"]),
            "source_range": "A3:H30 (titre fusionne en ligne 4)",
            "last_updated": datetime.now().isoformat(),
        }

    @staticmethod
    def get_fiche_appro_viande(file_path: Optional[Path] = None) -> Dict[str, Any]:
        """Extract Controle Appro Viande from A3:H3 headers, title in row 4, data rows 5..18."""
        config = EXCEL_FILES["fiche_appro_viande"]
        source_file = ExcelReader._resolve_file_path(config["file_name"], file_path)
        if not source_file.exists():
            raise FileNotFoundError(f"File not found: {source_file}")

        workbook = openpyxl.load_workbook(source_file, data_only=True)
        worksheet = ExcelReader._get_sheet(workbook, config)

        extracted = ExcelReader._extract_rect_table(
            worksheet,
            header_row=3,
            data_start_row=5,
            data_end_row=18,
            start_col=1,
            end_col=8,
            section_title_row=4,
        )

        workbook.close()
        return {
            "name": "Contrôle Appro Viande",
            "data": extracted["data"],
            "columns": extracted["columns"],
            "row_count": len(extracted["data"]),
            "source_range": "A3:H18 (titre fusionne en ligne 4)",
            "last_updated": datetime.now().isoformat(),
        }
    
    @staticmethod
    def get_all_data() -> Dict[str, Any]:
        """Load all Excel data"""
        return {
            "fiche_consommation": ExcelReader.get_fiche_consommation(),
            "fiche_emb_ingredient": ExcelReader.get_fiche_emb_ingredient(),
            "fiche_appro_viande": ExcelReader.get_fiche_appro_viande(),
            "calcul_viande": ExcelReader.get_calcul_viande(),
            "emballage_synthese": ExcelReader.get_emballage_synthese(),
            "timestamp": datetime.now().isoformat()
        }
